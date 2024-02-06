import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Border, Side


def get_schedule_course(codigo_course, trimestre):
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    horas = ['h6_7', 'h7_8', 'h8_9', 'h9_10', 'h10_11', 'h11_12', 'h12_13', 'h13_14', 'h14_15', 'h15_16', 'h16_17', 'h17_18', 'h18_19', 'h19_20', 'h20_21', 'h21_22']
    letras = ['L', 'M', 'MI', 'J', 'V', 'S']

    df = pd.read_excel(f'static/horarios/{trimestre}.xlsx')
    df = df.drop([0])
    df.columns = ['Source', 'FICHA', 'FORMACION', 'TITULAR', 'TRIMESTRE_ACADÉMICO', 'COMPETENCIA', 'NOMBRE_COMPETENCIA', 'RAP 1', 'RAP 2', 'RAP 3', 'RAP 4', 'RAP 5', 'RAP 6', 'INSTRUCTOR', 'NC2', 'HORAS_SEMANAL', 'NN2', 'h6_7L', 'h7_8L', 'h8_9L', 'h9_10L', 'h10_11L', 'h11_12L', 'h12_13L', 'h13_14L', 'h14_15L', 'h15_16L', 'h16_17L', 'h17_18L', 'h18_19L', 'h19_20L', 'h20_21L', 'h21_22L', 'h6_7M', 'h7_8M', 'h8_9M', 'h9_10M', 'h10_11M', 'h11_12M', 'h12_13M', 'h13_14M', 'h14_15M', 'h15_16M', 'h16_17M', 'h17_18M', 'h18_19M', 'h19_20M', 'h20_21M', 'h21_22M', 'h6_7MI', 'h7_8MI', 'h8_9MI', 'h9_10MI', 'h10_11MI', 'h11_12MI', 'h12_13MI', 'h13_14MI', 'h14_15MI', 'h15_16MI', 'h16_17MI', 'h17_18MI', 'h18_19MI', 'h19_20MI', 'h20_21MI', 'h21_22MI', 'h6_7J', 'h7_8J', 'h8_9J', 'h9_10J', 'h10_11J', 'h11_12J', 'h12_13J', 'h13_14J', 'h14_15J', 'h15_16J', 'h16_17J', 'h17_18J', 'h18_19J', 'h19_20J', 'h20_21J', 'h21_22J', 'h6_7V', 'h7_8V', 'h8_9V', 'h9_10V', 'h10_11V', 'h11_12V', 'h12_13V', 'h13_14V', 'h14_15V', 'h15_16V', 'h16_17V', 'h17_18V', 'h18_19V', 'h19_20V', 'h20_21V', 'h21_22V', 'h6_7S', 'h7_8S', 'h8_9S', 'h9_10S', 'h10_11S', 'h11_12S', 'h12_13S', 'h13_14S', 'h14_15S', 'h15_16S', 'h16_17S', 'h17_18S', 'h18_19S', 'h19_20S', 'h20_21S', 'h21_22S']
    
    df = df.drop(['Source', 'COMPETENCIA', 'RAP 1', 'RAP 2', 'RAP 3', 'RAP 4', 'RAP 5', 'RAP 6', 'NC2', 'HORAS_SEMANAL', 'NN2', 'HORAS_SEMANAL',], axis=1)
    

    df['NOMBRE_COMPETENCIA'] = df['NOMBRE_COMPETENCIA'].str.split('-').str[-1].str.strip()
    df['FICHA'] = df['FICHA'].str.split().str[0]
    df = df[df['FICHA'].str.strip() == codigo_course]

    schedule_course = {"ficha": codigo_course, "horario": {}}
    

    for dia, letra in zip(dias, letras):
        horario_dia = {}
        for hora in horas:
            columna = f"{hora}{letra}"
            if columna in df.columns:
                clases = df[df[columna].notnull()][[columna, 'NOMBRE_COMPETENCIA', 'INSTRUCTOR', 'FORMACION']].rename(columns={columna: 'AMB'}).to_dict('records')
                horario_dia[hora] = clases
        schedule_course["horario"][dia] = horario_dia

    df_itinerarios = pd.read_excel('static/Itinerarios.xlsx')    
    df_itinerarios.columns = ['NCL', 'COMPETENCIA', 'PALABRA_CLAVE', 'N_RAP', 'RAP', 'VERSION_ITINERARIO', 'RAP_SOFIA', 'PROGRAMA']    
    for dia in schedule_course['horario']:
        for hora in schedule_course['horario'][dia]:
            for clase in schedule_course['horario'][dia][hora]:
                clase['detalle'] = df_itinerarios[(df_itinerarios['PALABRA_CLAVE'] == clase['NOMBRE_COMPETENCIA']) & (df_itinerarios['PROGRAMA'] == clase['FORMACION'])].to_dict('records')

    detalles = {"key": []}
    detalles_set = set()
    for dia in schedule_course['horario']:
        for hora in schedule_course['horario'][dia]:
            for clase in schedule_course['horario'][dia][hora]:
                detalle = clase['detalle']
                if detalle:
                    detalle[0]['PALABRA_CLAVE'] = clase['NOMBRE_COMPETENCIA']        
                    detalle_tuple = tuple(detalle[0].items())
                    if detalle_tuple not in detalles_set:  # Check if the tuple is not in the set
                        detalles["key"].append(detalle)
                        detalles_set.add(detalle_tuple)
    
    schedule_course['detalles'] = detalles   

    schedule_course['titular'] = df['TITULAR'].iloc[0]

    schedule_course['trimestre'] = df['TRIMESTRE_ACADÉMICO'].iloc[0]

    return schedule_course


def generar_excel_course(schedule, nombre_archivo):    
    data = {'Hora': [], 'Lunes': [], 'Martes': [], 'Miércoles': [], 'Jueves': [], 'Viernes': [], 'Sábado': []}
    hours = ['6:00 - 7:00', '7:00 - 8:00', '8:00 - 9:00', '9:00 - 10:00', '10:00 - 11:00', '11:00 - 12:00', '12:00 - 13:00', '13:00 - 14:00', '14:00 - 15:00', '15:00 - 16:00', '16:00 - 17:00', '17:00 - 18:00', '18:00 - 19:00', '19:00 - 20:00', '20:00 - 21:00', '21:00 - 22:00']

    for hour in hours:
        data['Hora'].append(hour)  # Agregar la hora al DataFrame
        for dia, horarios in schedule['horario'].items():
            classes_info = []
            for hora, clases in horarios.items():
                if hora == f"h{hours.index(hour) + 6}_{hours.index(hour) + 7}":  # Comprobar si la hora coincide
                    for clase in clases:
                        # Obtener la información de FICHA, AMB y el nombre de la competencia
                        class_info = f"{clase['NOMBRE_COMPETENCIA']} \n {clase['INSTRUCTOR']} \n {clase['AMB']}"
                        classes_info.append(class_info)
                    break
            data[dia].append(", ".join(classes_info) if classes_info else '')  # Si no hay clases, añadir cadena vacía

    df = pd.DataFrame(data)
    
    workbook = Workbook()
    hoja = workbook.active   
    
    img = Image('static/images/logo-sena.png')
    hoja.add_image(img, f'A1')

    """ hoja.merge_cells('B2:C2')
    hoja.merge_cells('B3:C3') """

    hoja['B2'] = 'Centro de Servicios y Gestión Empresarial'
    hoja['B3'] = 'Coordinación de Teleinformática'
    hoja['A4'] = ''
    hoja['A5'] = 'Ficha:'
    hoja['B5'] = schedule['ficha']
    hoja['A6'] = 'Trimestre académico:'
    hoja['B6'] = '1-2024'
    hoja['A7'] = ''
    hoja['D5'] = 'Instructor titular:'
    hoja['E5'] = schedule['titular']
    hoja[f'A5'].font = hoja[f'A5'].font.copy(bold=True)
    hoja[f'A6'].font = hoja[f'A6'].font.copy(bold=True)
    hoja[f'D5'].font = hoja[f'D5'].font.copy(bold=True)



    for i in range(2, 4):
        hoja[f'B{i}'].font = hoja[f'C{i}'].font.copy(bold=True)
        """ hoja[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')  """  
    
       
    hoja.column_dimensions['A'].width = 19
    hoja.column_dimensions['B'].width = 25
    hoja.column_dimensions['C'].width = 25
    hoja.column_dimensions['D'].width = 25
    hoja.column_dimensions['E'].width = 25
    hoja.column_dimensions['F'].width = 25
    hoja.column_dimensions['G'].width = 25   

    for r in dataframe_to_rows(df, index=False, header=True):
        hoja.append(r)

    for i in range(8, 25):        
        hoja[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'D{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'E{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'F{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hoja[f'G{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rango_celdas = hoja['A8:G8']
    
    color_fondo = PatternFill(start_color='C5EAE8', end_color='C5EAE8', fill_type='solid')
    
    for row in rango_celdas:
        for cell in row:
            cell.fill = color_fondo
            cell.font = cell.font.copy(bold=True)

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    for row in hoja['A8:G25']:
        for cell in row:
            cell.border = thin_border
    
    for row in hoja['A8:G25']:
        for cell in row:
            cell.border = thin_border     
    
    hoja['A27'] = 'Palabra clave'
    hoja['B27'] = 'Norma de Competencia (Sofíaplus)'
    hoja['C27'] = 'Competencia'
    hoja['D27'] = 'RAP'
    details = schedule['detalles']
    
    row_num = 28
    for key, detail_list in details.items():
        for detail in detail_list:
            for item in detail:          
                
                hoja[f'A{row_num}'] = item['PALABRA_CLAVE']
                hoja[f'A{row_num}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                hoja[f'B{row_num}'] = item['NCL']
                hoja[f'B{row_num}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                hoja[f'C{row_num}'] = item['COMPETENCIA']
                hoja[f'C{row_num}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                hoja[f'D{row_num}'] = item['RAP']
                hoja[f'D{row_num}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                row_num += 1
    rango_celdas = hoja['A27:D27']
    
    color_fondo = PatternFill(start_color='C5EAE8', end_color='C5EAE8', fill_type='solid')
    
    for row in rango_celdas:
        for cell in row:
            cell.fill = color_fondo
            cell.font = cell.font.copy(bold=True)
        
    # Guardar el archivo Excel
    workbook.save(f'static/schedule-courses/{nombre_archivo}')


